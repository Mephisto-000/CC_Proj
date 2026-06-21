"""CommitHistoryWriter 測試。"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from daily_git_history.domain.models import CommitRecord, FileChange
from daily_git_history.services.xlsx_writer import CommitHistoryWriter


def _rows(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


class TestCommitHistoryWriter:
    def test_writes_header_row(self, tmp_path: Path) -> None:
        path = tmp_path / "20260622.xlsx"
        CommitHistoryWriter().write([], path)
        assert _rows(path) == [("作者", "Commit 訊息", "變動檔案")]

    def test_expands_multi_file_commit_into_multiple_rows_sorted_by_extension(
        self, tmp_path: Path
    ) -> None:
        record = CommitRecord(
            commit_hash="abc123",
            author="Mephisto",
            committed_at=datetime(2026, 6, 22, 10, 0, 0),
            subject="feat: 新增功能",
            files=(
                FileChange(path="b.py"),
                FileChange(path="README.md"),
                FileChange(path="a.py"),
            ),
        )
        path = tmp_path / "20260622.xlsx"
        CommitHistoryWriter().write([record], path)

        rows = _rows(path)
        assert rows[0] == ("作者", "Commit 訊息", "變動檔案")
        assert rows[1:] == [
            ("Mephisto", "feat: 新增功能", "README.md"),
            ("Mephisto", "feat: 新增功能", "a.py"),
            ("Mephisto", "feat: 新增功能", "b.py"),
        ]

    def test_commit_without_files_writes_blank_file_column(
        self, tmp_path: Path
    ) -> None:
        record = CommitRecord(
            commit_hash="abc123",
            author="Mephisto",
            committed_at=datetime(2026, 6, 22, 10, 0, 0),
            subject="merge",
            files=(),
        )
        path = tmp_path / "20260622.xlsx"
        CommitHistoryWriter().write([record], path)

        assert _rows(path)[1] == ("Mephisto", "merge", None)

    def test_preserves_commit_order_top_to_bottom(self, tmp_path: Path) -> None:
        earlier = CommitRecord(
            commit_hash="a",
            author="A",
            committed_at=datetime(2026, 6, 22, 9, 0, 0),
            subject="first",
            files=(FileChange(path="x.py"),),
        )
        later = CommitRecord(
            commit_hash="b",
            author="B",
            committed_at=datetime(2026, 6, 22, 18, 0, 0),
            subject="second",
            files=(FileChange(path="y.py"),),
        )
        path = tmp_path / "20260622.xlsx"
        CommitHistoryWriter().write([earlier, later], path)

        rows = _rows(path)
        assert rows[1][1] == "first"
        assert rows[2][1] == "second"

    def test_creates_parent_directories(self, tmp_path: Path) -> None:
        path = tmp_path / "nested" / "dir" / "20260622.xlsx"
        CommitHistoryWriter().write([], path)
        assert path.exists()
