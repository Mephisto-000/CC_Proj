"""daily_git_history.cli 測試。

包含使用 fake reader/writer 的單元測試，以及對真實 tmp git repository
執行完整流程的整合測試。
"""

import subprocess
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from daily_git_history import cli
from daily_git_history.domain.exceptions import GitCommandError
from daily_git_history.domain.models import CommitRecord, FileChange
from daily_git_history.services.git_log_reader import DEFAULT_TIMEZONE


class _FakeReader:
    def __init__(self, records: list[CommitRecord]) -> None:
        self._records = records
        self.requested_date: date | None = None

    def read(self, target_date: date) -> list[CommitRecord]:
        self.requested_date = target_date
        return self._records


class _FakeWriter:
    def __init__(self) -> None:
        self.written: tuple[list[CommitRecord], Path] | None = None

    def write(self, records: list[CommitRecord], path: Path) -> None:
        self.written = (records, path)


class _FailingReader:
    def read(self, target_date: date) -> list[CommitRecord]:
        raise GitCommandError("模擬失敗")


class TestBuildOutputPath:
    def test_filename_is_eight_digit_date(self, tmp_path: Path) -> None:
        path = cli.build_output_path(date(2026, 6, 22), tmp_path)
        assert path == tmp_path / "20260622.xlsx"


class TestParseArgs:
    def test_defaults(self) -> None:
        args = cli.parse_args([])
        assert args.repo == Path(".")
        assert args.output_dir == cli.DEFAULT_OUTPUT_DIR
        assert args.date is None

    def test_overrides(self) -> None:
        args = cli.parse_args(
            ["--repo", "/tmp/repo", "--output-dir", "/tmp/out", "--date", "2026-06-22"]
        )
        assert args.repo == Path("/tmp/repo")
        assert args.output_dir == Path("/tmp/out")
        assert args.date == date(2026, 6, 22)


class TestResolveTargetDate:
    def test_returns_explicit_date_when_provided(self) -> None:
        assert cli.resolve_target_date(date(2026, 1, 1)) == date(2026, 1, 1)

    def test_defaults_to_taipei_today_when_not_provided(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        # 固定「現在」為台灣時間 2026-06-22 00:30，驗證即使伺服器在
        # UTC 時區執行（當地仍是 06-21），仍以台灣日期為準。
        fixed_now = datetime(2026, 6, 22, 0, 30, tzinfo=DEFAULT_TIMEZONE)

        class _FixedDateTime(datetime):
            @classmethod
            def now(cls, tz=None):  # type: ignore[override]
                return fixed_now

        monkeypatch.setattr(cli, "datetime", _FixedDateTime)
        assert cli.resolve_target_date(None) == date(2026, 6, 22)


class TestRun:
    def test_uses_injected_reader_and_writer(self, tmp_path: Path) -> None:
        record = CommitRecord(
            commit_hash="abc",
            author="Mephisto",
            committed_at=datetime(2026, 6, 22, 10, 0, 0),
            subject="feat: 範例",
            files=(FileChange(path="a.py"),),
        )
        reader = _FakeReader([record])
        writer = _FakeWriter()

        output_path = cli.run(
            repo_path=tmp_path,
            output_dir=tmp_path / "out",
            target_date=date(2026, 6, 22),
            reader=reader,
            writer=writer,
        )

        assert reader.requested_date == date(2026, 6, 22)
        assert output_path == tmp_path / "out" / "20260622.xlsx"
        assert writer.written is not None
        assert writer.written[0] == [record]
        assert writer.written[1] == output_path

    def test_propagates_reader_errors(self, tmp_path: Path) -> None:
        with pytest.raises(GitCommandError):
            cli.run(
                repo_path=tmp_path,
                output_dir=tmp_path,
                target_date=date(2026, 6, 22),
                reader=_FailingReader(),
                writer=_FakeWriter(),
            )


def _init_repo_with_commits(repo_dir: Path) -> None:
    def _git(*args: str, env: dict[str, str] | None = None) -> None:
        subprocess.run(
            ["git", "-C", str(repo_dir), *args],
            check=True,
            capture_output=True,
            env=env,
        )

    import os

    _git("init", "-q")
    _git("config", "user.name", "Mephisto")
    _git("config", "user.email", "jackorgordon@gmail.com")

    (repo_dir / "a.py").write_text("print('a')\n", encoding="utf-8")
    (repo_dir / "README.md").write_text("# demo\n", encoding="utf-8")
    _git("add", "a.py", "README.md")
    env_first = {
        **os.environ,
        "GIT_AUTHOR_DATE": "2026-06-22T09:00:00+08:00",
        "GIT_COMMITTER_DATE": "2026-06-22T09:00:00+08:00",
    }
    _git("commit", "-q", "-m", "feat: 初始化專案", env=env_first)

    (repo_dir / "b.py").write_text("print('b')\n", encoding="utf-8")
    (repo_dir / ".gitignore").write_text("*.pyc\n", encoding="utf-8")
    _git("add", "b.py", ".gitignore")
    env_second = {
        **os.environ,
        "GIT_AUTHOR_DATE": "2026-06-22T18:00:00+08:00",
        "GIT_COMMITTER_DATE": "2026-06-22T18:00:00+08:00",
    }
    _git("commit", "-q", "-m", "feat: 新增 b 模組", env=env_second)


class TestMainIntegration:
    def test_generates_xlsx_from_real_repo(self, tmp_path: Path) -> None:
        repo_dir = tmp_path / "repo"
        repo_dir.mkdir()
        _init_repo_with_commits(repo_dir)

        output_dir = tmp_path / "DailyGitHistory"
        cli.main(
            [
                "--repo",
                str(repo_dir),
                "--output-dir",
                str(output_dir),
                "--date",
                "2026-06-22",
            ]
        )

        output_path = output_dir / "20260622.xlsx"
        assert output_path.exists()

        workbook = load_workbook(output_path)
        sheet = workbook.active
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]

        assert rows[0] == ("作者", "Commit 訊息", "變動檔案")
        # 第一個 commit（09:00）在上，依副檔名排序：README.md 在 a.py 之前。
        assert rows[1] == ("Mephisto", "feat: 初始化專案", "README.md")
        assert rows[2] == ("Mephisto", "feat: 初始化專案", "a.py")
        # 第二個 commit（18:00）在下，.gitignore（無副檔名）排在 b.py 之前。
        assert rows[3] == ("Mephisto", "feat: 新增 b 模組", ".gitignore")
        assert rows[4] == ("Mephisto", "feat: 新增 b 模組", "b.py")

    def test_main_raises_on_invalid_repo(self, tmp_path: Path) -> None:
        not_a_repo = tmp_path / "not-a-repo"
        not_a_repo.mkdir()
        with pytest.raises(GitCommandError):
            cli.main(["--repo", str(not_a_repo), "--output-dir", str(tmp_path / "out")])
